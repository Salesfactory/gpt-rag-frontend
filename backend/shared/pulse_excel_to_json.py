"""
Hey team:
When this script is was designed, only god and me knew how it worked

Now, probably only god knows it!

Therefore, if you are trying to make any improvements and it fails parsing (most surely)
Please increase this counter as a warning for the next person 

total_hours_waster_here: 8
"""

from __future__ import annotations

import json
import re
from collections import defaultdict
from dataclasses import dataclass, field
from enum import Enum, auto
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


__all__ = [
    "serialize_excel",
    "serialize_to_file",
    "ExcelParserError",
]


class ExcelParserError(Exception):
    pass


class UnsupportedFileFormatError(ExcelParserError):
    pass


class NoValidSheetsError(ExcelParserError):
    pass


VALID_EXTENSIONS = (".xlsx",)
TARGET_SHEET_NAMES = frozenset({"Full Run %", "%"})
MINIMUM_HEADER_ROWS = 2

CATEGORY_COL = 0
N_COL = 1
PERCENT_COL = 2

HEADER_ROW_1 = 0
HEADER_ROW_2 = 1
ROW_TO_DELETE = 2
DATA_START_ROW = 2


@dataclass
class ColumnGroup:
    name: str
    key: str
    start_col: int
    sub_columns: dict[int, str] = field(default_factory=dict)


class RowType(Enum):
    SECTION_HEADER = auto()
    PARENT_CATEGORY = auto()
    DATA_ROW = auto()
    EMPTY = auto()


@dataclass
class ParsedRow:
    row_type: RowType
    category: str | None = None
    values: dict[int, float] = field(default_factory=dict)


@dataclass(frozen=True)
class RowContext:
    prev_empty: bool
    next_empty: bool
    body_empty: bool


@dataclass
class DataRecord:
    section: str
    category: str
    parent_category: str
    breakdowns: dict[str, dict[str, float]] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        result: dict[str, Any] = {
            "section": self.section,
            "category": self.category,
            "parent_category": self.parent_category,
        }
        result.update(self.breakdowns)
        return result


def _to_snake_case(name: str) -> str:
    result = re.sub(r"[^a-zA-Z0-9]+", "_", name.lower()).strip("_")
    return f"by_{result}"


def _is_numeric(value: Any) -> bool:
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        try:
            float(value)
            return True
        except ValueError:
            return False
    return False


def _is_all_caps(text: str) -> bool:
    letters = [c for c in text if c.isalpha()]
    return len(letters) > 0 and all(c.isupper() for c in letters)


def _cell_is_empty(cell: Any) -> bool:
    return cell is None or str(cell).strip() == ""


def _row_is_empty(row: list[Any]) -> bool:
    return all(_cell_is_empty(cell) for cell in row)


def _format_category_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(int(value) if value == int(value) else value)
    stripped = str(value).strip()
    return stripped if stripped else None


class ColumnMapper:
    def __init__(self, row1: list[Any], row2: list[Any]) -> None:
        self._column_groups: list[ColumnGroup] = []
        self._parse_headers(row1, row2)

    @property
    def column_groups(self) -> list[ColumnGroup]:
        return self._column_groups

    def _parse_headers(self, row1: list[Any], row2: list[Any]) -> None:
        current_group: ColumnGroup | None = None
        max_cols = max(len(row1), len(row2))

        for col_idx in range(max_cols):
            row1_val = row1[col_idx] if col_idx < len(row1) else None
            row2_val = row2[col_idx] if col_idx < len(row2) else None

            if col_idx == CATEGORY_COL:
                continue

            if col_idx == N_COL and self._is_n_column(row2_val):
                current_group = self._create_total_group(col_idx)
                self._column_groups.append(current_group)
                continue

            if col_idx == PERCENT_COL and self._is_percent_column(row2_val, current_group):
                current_group.sub_columns[col_idx] = "%"
                continue

            if self._has_value(row1_val):
                current_group = self._create_column_group(row1_val, col_idx)
                self._column_groups.append(current_group)

            if self._has_value(row2_val) and current_group is not None:
                current_group.sub_columns[col_idx] = str(row2_val).strip()

    def _is_n_column(self, value: Any) -> bool:
        return value is not None and str(value).strip().upper() == "N"

    def _is_percent_column(self, value: Any, current_group: ColumnGroup | None) -> bool:
        return (
            value is not None
            and str(value).strip() == "%"
            and current_group is not None
            and current_group.key == "by_total"
        )

    def _has_value(self, value: Any) -> bool:
        return value is not None and str(value).strip() != ""

    def _create_total_group(self, col_idx: int) -> ColumnGroup:
        group = ColumnGroup(name="Total", key="by_total", start_col=col_idx)
        group.sub_columns[col_idx] = "N"
        return group

    def _create_column_group(self, name: Any, col_idx: int) -> ColumnGroup:
        group_name = str(name).strip()
        return ColumnGroup(
            name=group_name,
            key=_to_snake_case(group_name),
            start_col=col_idx,
        )


class RowParser:
    def parse(self, row: list[Any]) -> ParsedRow:
        if not row or _row_is_empty(row):
            return ParsedRow(row_type=RowType.EMPTY)

        category_val = row[0]
        data_columns = row[1:] if len(row) > 1 else []

        values = self._extract_numeric_values(data_columns)
        has_data = bool(values)

        category = _format_category_value(category_val)
        is_numeric_category = isinstance(category_val, (int, float))

        return self._classify_row(category, has_data, is_numeric_category, values)

    def _extract_numeric_values(self, columns: list[Any]) -> dict[int, float]:
        values: dict[int, float] = {}
        for col_idx, cell in enumerate(columns, start=1):
            if cell is not None and _is_numeric(cell):
                values[col_idx] = float(cell)
        return values

    def _classify_row(
        self,
        category: str | None,
        has_data: bool,
        is_numeric_category: bool,
        values: dict[int, float],
    ) -> ParsedRow:
        if has_data:
            return ParsedRow(row_type=RowType.DATA_ROW, category=category, values=values)

        if not category:
            return ParsedRow(row_type=RowType.EMPTY)

        if is_numeric_category:
            return ParsedRow(row_type=RowType.DATA_ROW, category=category, values=values)

        if _is_all_caps(category):
            return ParsedRow(row_type=RowType.SECTION_HEADER, category=category)

        return ParsedRow(row_type=RowType.PARENT_CATEGORY, category=category)


class RecordBuilder:
    def __init__(self, column_mapper: ColumnMapper) -> None:
        self._column_mapper = column_mapper

    def build(
        self,
        parsed_row: ParsedRow,
        current_section: str,
        current_parent: str,
    ) -> DataRecord:
        record = DataRecord(
            section=current_section,
            category=parsed_row.category or "",
            parent_category=current_parent,
        )

        for group in self._column_mapper.column_groups:
            breakdown = self._extract_breakdown(group, parsed_row.values)
            if breakdown:
                record.breakdowns[group.key] = breakdown

        return record

    def _extract_breakdown(self, group: ColumnGroup, values: dict[int, float]) -> dict[str, float]:
        return {
            sub_col_name: values[col_idx]
            for col_idx, sub_col_name in group.sub_columns.items()
            if col_idx in values
        }


class ExcelParser:
    def __init__(self, file_source: str | Path | BytesIO) -> None:
        self._file_source = file_source
        self._workbook: Workbook | None = None
        self._load_file()

    @property
    def sheet_names(self) -> list[str]:
        if self._workbook is None:
            return []
        return self._workbook.sheetnames

    def _load_file(self) -> None:
        if isinstance(self._file_source, BytesIO):
            try:
                self._workbook = openpyxl.load_workbook(self._file_source, data_only=True)
            except Exception as e:
                raise ExcelParserError(f"Unable to parse Excel file: {e}") from e
            return

        file_path = Path(self._file_source)
        if file_path.suffix.lower() not in VALID_EXTENSIONS:
            raise UnsupportedFileFormatError(
                f"Unsupported file format: {file_path.suffix}. "
                f"Expected one of: {', '.join(VALID_EXTENSIONS)}"
            )

        try:
            self._workbook = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            raise ExcelParserError(f"Unable to parse Excel file: {e}") from e

    def get_sheet_data(self, sheet_name: str) -> list[list[Any]]:
        if self._workbook is None:
            return []
        sheet: Worksheet = self._workbook[sheet_name]
        return [[cell.value for cell in row] for row in sheet.iter_rows()]


class SheetProcessor:
    def __init__(self, sheet_data: list[list[Any]]) -> None:
        self._raw_data = sheet_data
        self._cleaned_data: list[list[Any]] | None = None
        self._column_mapper: ColumnMapper | None = None
        self._row_parser: RowParser | None = None
        self._record_builder: RecordBuilder | None = None

    def process(self) -> list[dict[str, Any]]:
        if len(self._raw_data) < MINIMUM_HEADER_ROWS:
            return []

        self._cleaned_data = self._prepare_data()
        self._initialize_components()
        return self._extract_records()

    def _prepare_data(self) -> list[list[Any]]:
        data = [row[:] for row in self._raw_data]
        del data[ROW_TO_DELETE]
        data = self._clean_orphan_values(data)

        if data[HEADER_ROW_1] and _cell_is_empty(data[HEADER_ROW_1][CATEGORY_COL]):
            data[HEADER_ROW_1][CATEGORY_COL] = "Category"

        return data

    def _clean_orphan_values(self, data: list[list[Any]]) -> list[list[Any]]:
        for row in data:
            non_empty = [
                (i, cell)
                for i, cell in enumerate(row)
                if not _cell_is_empty(cell)
            ]
            if len(non_empty) == 1 and non_empty[0][0] != CATEGORY_COL:
                row[non_empty[0][0]] = None
        return data

    def _initialize_components(self) -> None:
        row1 = self._cleaned_data[HEADER_ROW_1]
        row2 = self._cleaned_data[HEADER_ROW_2]
        self._column_mapper = ColumnMapper(row1, row2)
        self._row_parser = RowParser()
        self._record_builder = RecordBuilder(self._column_mapper)

    def _extract_records(self) -> list[dict[str, Any]]:
        data_rows = self._cleaned_data[DATA_START_ROW:]
        parsed_rows = [self._row_parser.parse(row) for row in data_rows]

        current_section = ""
        current_parent = ""
        records: list[dict[str, Any]] = []

        for i, parsed_row in enumerate(parsed_rows):
            if parsed_row.row_type == RowType.EMPTY:
                continue

            context = self._get_row_context(i, parsed_rows, data_rows)

            if parsed_row.row_type in (RowType.SECTION_HEADER, RowType.PARENT_CATEGORY):
                current_section, current_parent = self._update_hierarchy(
                    parsed_row, context, current_section, current_parent
                )
            elif parsed_row.row_type == RowType.DATA_ROW:
                record = self._record_builder.build(parsed_row, current_section, current_parent)
                records.append(record.to_dict())

        return records

    def _get_row_context(
        self,
        index: int,
        parsed_rows: list[ParsedRow],
        data_rows: list[list[Any]],
    ) -> RowContext:
        prev_empty = index == 0 or parsed_rows[index - 1].row_type == RowType.EMPTY
        next_empty = index == len(parsed_rows) - 1 or parsed_rows[index + 1].row_type == RowType.EMPTY

        current_row = data_rows[index]
        body_cells = current_row[CATEGORY_COL + 1:]
        body_empty = all(_cell_is_empty(cell) for cell in body_cells) if body_cells else True

        return RowContext(prev_empty=prev_empty, next_empty=next_empty, body_empty=body_empty)

    def _update_hierarchy(
        self,
        parsed_row: ParsedRow,
        context: RowContext,
        current_section: str,
        current_parent: str,
    ) -> tuple[str, str]:
        category = parsed_row.category or ""

        if context.prev_empty and context.body_empty and not context.next_empty:
            return category, current_parent

        if context.prev_empty and context.body_empty and context.next_empty:
            return current_section, category

        return current_section, current_parent


def _serialize_to_rows(file_source: str | Path | BytesIO) -> list[dict[str, Any]]:
    parser = ExcelParser(file_source)
    sheets_to_process = [s for s in parser.sheet_names if s in TARGET_SHEET_NAMES]

    if not sheets_to_process:
        raise NoValidSheetsError(
            f"No valid sheets found. Expected one of {sorted(TARGET_SHEET_NAMES)}, "
            f"but found: {parser.sheet_names}"
        )

    all_records: list[dict[str, Any]] = []

    for sheet_name in sheets_to_process:
        sheet_data = parser.get_sheet_data(sheet_name)
        processor = SheetProcessor(sheet_data)
        all_records.extend(processor.process())

    return all_records


def serialize_excel(file_source: str | Path | BytesIO) -> list[dict[str, Any]]:
    row_records = _serialize_to_rows(file_source)
    return _group_by_column(row_records)


def serialize_to_file(input_path: str | Path, output_path: str | Path) -> None:
    records = serialize_excel(input_path)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)


def _group_by_column(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not records:
        return []

    by_section_parent: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    by_columns: set[str] = set()

    for record in records:
        key = (record.get("section", ""), record.get("parent_category", ""))
        by_section_parent[key].append(record)
        by_columns.update(k for k in record if k.startswith("by_"))

    result: list[dict[str, Any]] = []

    for (section, parent_category), group_records in by_section_parent.items():
        for column in sorted(by_columns):
            data: dict[str, dict[str, float]] = defaultdict(dict)

            for record in group_records:
                category = record.get("category", "Unknown")
                column_data = record.get(column, {})

                if not isinstance(column_data, dict):
                    continue

                for segment, value in column_data.items():
                    data[segment][category] = round(value, 2) if isinstance(value, float) else value

            if data:
                result.append({
                    "section": section,
                    "parent_category": parent_category,
                    "column": column,
                    "data": dict(data),
                })

    return result


if __name__ == "__main__":
    input_file = "Pulse 145 Databook.xlsx"
    output_file = Path(input_file).with_suffix(".json")
    serialize_to_file(input_file, output_file)
